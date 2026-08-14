"""
main.py - Script Único de Automação de Criação de Títulos no Senior ERP
========================================================================
Este script centraliza toda a lógica de automação:
1. Leitura do arquivo config/config.json (coordenadas, mapeamento das filiais e delays).
2. Leitura da fila de títulos tratados em data/titulos_tratados.json.
3. Contagem regressiva de 5 segundos.
4. Troca de filial inteligente (apenas quando a filial muda).
5. Confirmação condicional de fechar telas ao trocar de filial.
6. Sequência de cliques nos menus (maleta > finanças > contas_receber > entrada_manutencao > titulos_manutencao).
7. Preenchimento automático dos campos no formulário F301TCR até "Data Entrada".
8. Fail-Safe: Se o mouse for movido para o canto superior esquerdo (X <= 15, Y <= 15), a automação para IMEDIATAMENTE.
"""

import os
import json
import time
import sys
import csv
from datetime import datetime
from pathlib import Path
import pyautogui
import pyperclip

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_DISPONIVEL = True
except ImportError:
    EXCEL_DISPONIVEL = False

# Habilitar o recurso nativo de Fail-Safe do PyAutoGUI
pyautogui.FAILSAFE = True

# Diretores base do projeto
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
CONFIG_DIR = BASE_DIR / "config"
SCRIPTS_DIR = BASE_DIR / "scripts"
OUTPUT_DIR = BASE_DIR / "outputs"

# Garantir que a pasta outputs exista
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Arquivos de dados, configuração e outputs
TITULOS_JSON_PATH = DATA_DIR / "titulos_tratados.json"
CONFIG_JSON_PATH = CONFIG_DIR / "config.json"
PROGRESSO_JSON_PATH = OUTPUT_DIR / "progresso.json"

_hoje_str = datetime.now().strftime("%Y%m%d")
LOG_TXT_PATH = OUTPUT_DIR / f"log_execucao_{_hoje_str}.txt"
EXCEL_OUTPUT_PATH = OUTPUT_DIR / f"titulos_processados_{_hoje_str}.xlsx"
CSV_OUTPUT_PATH = OUTPUT_DIR / f"titulos_processados_{_hoje_str}.csv"

PULADOS_EXCEL_PATH = OUTPUT_DIR / f"titulos_pulados_{_hoje_str}.xlsx"
PULADOS_CSV_PATH   = OUTPUT_DIR / f"titulos_pulados_{_hoje_str}.csv"

# SISTEMA DE LOG
def log(mensagem, nivel="INFO"):
    """Imprime no console e registra no arquivo de log dentro da pasta outputs."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    texto_formatado = f"[{timestamp}] [{nivel}] {mensagem}"
    print(mensagem if nivel == "INFO" else f"[{nivel}] {mensagem}")
    try:
        with open(LOG_TXT_PATH, "a", encoding="utf-8") as f:
            f.write(texto_formatado + "\n")
    except Exception as e:
        print(f"[ERRO LOG] Não foi possível escrever no arquivo de log: {e}")

# SISTEMA DE PROGRESSO / CHECKPOINT
def obter_ultimo_indice_processado():
    """Lê o arquivo progresso.json na pasta outputs para saber de onde retomar."""
    if PROGRESSO_JSON_PATH.exists():
        try:
            with open(PROGRESSO_JSON_PATH, "r", encoding="utf-8") as f:
                dados = json.load(f)
                return dados.get("ultimo_indice_concluido", 0)
        except Exception as e:
            log(f"Erro ao ler arquivo de progresso: {e}. Iniciando do zero.", "AVISO")
    return 0

def salvar_progresso(indice_concluido, num_titulo=""):
    """Salva o índice do último título processado com sucesso."""
    try:
        dados = {
            "ultimo_indice_concluido": indice_concluido,
            "ultimo_titulo_processado": num_titulo,
            "ultima_atualizacao": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        with open(PROGRESSO_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log(f"Erro ao salvar progresso: {e}", "ERRO")

def limpar_progresso():
    """Remove o arquivo de progresso quando todos os títulos são concluídos."""
    if PROGRESSO_JSON_PATH.exists():
        try:
            PROGRESSO_JSON_PATH.unlink()
            log("Arquivo de progresso limpo (todos os títulos foram concluídos).")
        except Exception as e:
            log(f"Erro ao remover arquivo de progresso: {e}", "AVISO")

# SISTEMA DE PLANILHA (EXCEL / CSV)
def registrar_titulo_planilha(titulo, indice, status="SUCESSO"):
    """Registra o título processado na planilha (Excel ou CSV) dentro de outputs."""
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    num_titulo = titulo.get("numero_titulo", "")
    cliente = titulo.get("cliente", "")
    filial = str(titulo.get("projeto", "")).strip()
    valor = titulo.get("valor_original", "")
    vencimento = titulo.get("data_vencimento", "")
    obs = titulo.get("observacao", "")
    centro_custo = titulo.get("centro_custo", "")

    # Registrar em Excel se openpyxl estiver disponível
    if EXCEL_DISPONIVEL:
        try:
            if not EXCEL_OUTPUT_PATH.exists():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Baixas Processadas"
                headers = ["Item", "Nº Título", "Cliente", "Filial", "Valor (R$)", "Vencimento", "Centro Custo", "Observação", "Status", "Data/Hora"]
                ws.append(headers)
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            else:
                wb = openpyxl.load_workbook(EXCEL_OUTPUT_PATH)
                ws = wb.active

            row = [indice, num_titulo, cliente, filial, valor, vencimento, centro_custo, obs, status, agora]
            ws.append(row)
            wb.save(EXCEL_OUTPUT_PATH)
        except Exception as e:
            log(f"Erro ao gravar no Excel: {e}", "ERRO")

    # Registrar sempre em CSV como backup universal
    try:
        arquivo_existe = CSV_OUTPUT_PATH.exists()
        with open(CSV_OUTPUT_PATH, "a", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            if not arquivo_existe:
                writer.writerow(["Item", "Nº Título", "Cliente", "Filial", "Valor (R$)", "Vencimento", "Centro Custo", "Observação", "Status", "Data/Hora"])
            writer.writerow([indice, num_titulo, cliente, filial, valor, vencimento, centro_custo, obs, status, agora])
    except Exception as e:
        log(f"Erro ao gravar no CSV: {e}", "ERRO")

def registrar_titulo_pulado(titulo, indice, motivo="Cliente Duplicado"):
    """Registra um título que foi pulado em uma planilha separada 'titulos_pulados'."""
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    num_titulo = titulo.get("numero_titulo", "")
    cliente = titulo.get("cliente", "")
    filial = str(titulo.get("projeto", "")).strip()
    valor = titulo.get("valor_original", "")
    vencimento = titulo.get("data_vencimento", "")
    obs = titulo.get("observacao", "")

    if EXCEL_DISPONIVEL:
        try:
            if not PULADOS_EXCEL_PATH.exists():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Títulos Pulados"
                headers = ["Item", "Nº Título", "Cliente", "Filial", "Valor (R$)", "Vencimento", "Observação", "Motivo", "Data/Hora"]
                ws.append(headers)
                header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            else:
                wb = openpyxl.load_workbook(PULADOS_EXCEL_PATH)
                ws = wb.active

            ws.append([indice, num_titulo, cliente, filial, valor, vencimento, obs, motivo, agora])
            wb.save(PULADOS_EXCEL_PATH)
        except Exception as e:
            log(f"Erro ao gravar no Excel de pulados: {e}", "ERRO")

    try:
        arquivo_existe = PULADOS_CSV_PATH.exists()
        with open(PULADOS_CSV_PATH, "a", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            if not arquivo_existe:
                writer.writerow(["Item", "Nº Título", "Cliente", "Filial", "Valor (R$)", "Vencimento", "Observação", "Motivo", "Data/Hora"])
            writer.writerow([indice, num_titulo, cliente, filial, valor, vencimento, obs, motivo, agora])
    except Exception as e:
        log(f"Erro ao gravar no CSV de pulados: {e}", "ERRO")

def verificar_cliente_duplicado(nome_cliente):
    """
    Verifica se a grade de pesquisa do ERP possui mais de 1 cliente com o mesmo nome.
    Captura uma imagem (screenshot) da região a partir de X: 38, Y: 271 (60px de largura por 50px de altura)
    e identifica se existem pixels pretos/escuros (texto indicando a presença da 2ª linha de cliente).
    """
    try:
        aguardar_com_failsafe(0.5)

        # Screenshot da região da 2ª linha na grade de resultados do ERP
        img = pyautogui.screenshot(region=(38, 271, 60, 50))

        pixels_escuros = 0
        largura, altura = img.size

        for x in range(largura):
            for y in range(altura):
                r, g, b = img.getpixel((x, y))[:3]
                # Pixels escuros correspondentes ao texto (número/nome do cliente)
                if r < 100 and g < 100 and b < 100:
                    pixels_escuros += 1

        log(f"  [Verificação Imagem] Pixels escuros encontrados na região (X:38, Y:271): {pixels_escuros}")

        # Se encontrou pixels escuros suficientes, existe texto na 2ª linha (duplicidade)
        tem_duplicado = pixels_escuros > 15
        if tem_duplicado:
            log(f"  ➜ [CLIENTE DUPLICADO DETECTADO] Segunda linha de cliente encontrada para '{nome_cliente}'!", "AVISO")
        return tem_duplicado
    except Exception as e:
        log(f"  [Verificação Imagem] Erro ao analisar pixels da tela: {e}", "AVISO")
        return False


# Carregamento de Configurações
def carregar_config():
    if not CONFIG_JSON_PATH.exists():
        raise FileNotFoundError(f"Arquivo de configuração não encontrado em: {CONFIG_JSON_PATH}")
    with open(CONFIG_JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

CONFIG_DATA = carregar_config()
MAPA_FILIAIS = CONFIG_DATA.get("filiais", {})
BOTOES = CONFIG_DATA.get("botoes", {})
DELAYS = CONFIG_DATA.get("delays", {
    "home": 1.0,
    "key": 0.3,
    "enter": 1.5,
    "click": 1.0,
    "passo_menu": 1.0,
    "abrir_janela": 2.0,
    "entre_titulos": 1.5,
    "preenchimento": 1.0
})

# Carregamento de Dados
def carregar_titulos():
    if not TITULOS_JSON_PATH.exists():
        raise FileNotFoundError(f"Arquivo {TITULOS_JSON_PATH} não encontrado. Execute o script processar_planilha.py primeiro.")
    with open(TITULOS_JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

# Funções Auxiliares de Configuração
def obter_info_filial(codigo):
    codigo_str = str(codigo).strip()
    if codigo_str not in MAPA_FILIAIS:
        raise ValueError(f"Código de filial '{codigo}' não encontrado no config.json.")
    return MAPA_FILIAIS[codigo_str]

def obter_coordenadas_botao(nome_botao):
    nome_key = str(nome_botao).lower().strip()
    if nome_key not in BOTOES:
        raise ValueError(f"Botão '{nome_botao}' não encontrado no config.json.")
    return BOTOES[nome_key]

# Formatador de Data (YYYY-MM-DD -> DDMMYYYY)
def formatar_data_br(data_str, com_barras=False, inverter_dia_mes=False):
    """Converte datas do formato 'YYYY-MM-DD' para 'DDMMYYYY' ou 'DD/MM/YYYY'.
    Use inverter_dia_mes=True quando o ERP troca dia e mês automaticamente (digita MMDD em vez de DDMM).
    """
    if not data_str:
        return ""
    partes = str(data_str).strip().split("-")
    if len(partes) == 3:
        ano, mes, dia = partes[0], partes[1], partes[2]
        if inverter_dia_mes:
            # Inverte: digita mês primeiro para compensar o comportamento do ERP
            return f"{mes}/{dia}/{ano}" if com_barras else f"{mes}{dia}{ano}"
        return f"{dia}/{mes}/{ano}" if com_barras else f"{dia}{mes}{ano}"
    return str(data_str)

# Verificação contínua do Fail-Safe (canto superior esquerdo)
def verificar_failsafe():
    x, y = pyautogui.position()
    if x <= 15 and y <= 15:
        raise pyautogui.FailSafeException("Fail-Safe ativado: Mouse posicionado no canto superior esquerdo.")

def aguardar_com_failsafe(tempo_segundos):
    """Substitui o time.sleep para verificar o canto superior esquerdo durante a pausa."""
    passo = 0.05
    tempo_passado = 0.0
    while tempo_passado < tempo_segundos:
        verificar_failsafe()
        time.sleep(passo)
        tempo_passado += passo

# Funções de Automação do ERP
def iniciar_contagem_regressiva(segundos=5):
    print(f"\n=======================================================")
    print(f"[Automação] A automação iniciará em {segundos} segundos.")
    print(f"[Automação] Mude o foco para a janela do Senior ERP agora!")
    print(f"[DICA DE EMERGÊNCIA] Mova o mouse para o CANTO SUPERIOR ESQUERDO para PARAR a qualquer momento.")
    print(f"=======================================================")
    for i in range(segundos, 0, -1):
        verificar_failsafe()
        print(f"  >>> Iniciando em {i} segundo(s)...", end="\r", flush=True)
        aguardar_com_failsafe(1.0)
    print("\n\n[Automação] >>> INICIANDO AUTOMAÇÃO AGORA! <<<\n")

def digitar_texto(texto):
    """Digita texto usando clipboard para suportar caracteres especiais (ç, ã, á, etc.)."""
    pyperclip.copy(str(texto))
    pyautogui.hotkey("ctrl", "v")

def clicar_botao(nome_botao, double_click=False):
    verificar_failsafe()
    info_btn = obter_coordenadas_botao(nome_botao)
    x, y = info_btn["x"], info_btn["y"]
    desc = info_btn.get("descricao", nome_botao)

    if double_click:
        print(f"  [Automação] Clicando DUAS VEZES no botão '{desc}' em (X: {x}, Y: {y})...")
        pyautogui.doubleClick(x, y)
    else:
        print(f"  [Automação] Clicando no botão '{desc}' em (X: {x}, Y: {y})...")
        pyautogui.click(x, y)
        
    aguardar_com_failsafe(DELAYS.get("click", 0.8))

def abrir_tela_trocar_filial(tem_tela_aberta=False):
    print("  [Automação] Passo 1: Clicando no botão de Usuário...")
    clicar_botao("user")
    aguardar_com_failsafe(DELAYS.get("click", 0.8))

    print("  [Automação] Passo 2: Clicando no botão de Trocar Filial...")
    clicar_botao("trocar_filial")
    aguardar_com_failsafe(1.0)

    if tem_tela_aberta:
        print("  [Automação] Notificação de confirmação detectada. Confirmando fechamento de telas (Enter / Sim)...")
        verificar_failsafe()
        pyautogui.press("enter")
        aguardar_com_failsafe(DELAYS.get("abrir_janela", 2.0))
    else:
        print("  [Automação] Nenhuma tela aberta no ERP. Abrindo seleção de filial diretamente...")
        aguardar_com_failsafe(DELAYS.get("abrir_janela", 1.5))

def selecionar_filial(codigo, pressionar_enter=True):
    codigo_str = str(codigo).strip()
    info = obter_info_filial(codigo_str)
    passos = info["posicao"]
    fantasia = info.get("fantasia", "")

    print(f"  [Automação] Passo 3: Selecionando Filial {codigo_str} - {fantasia} (Posição {passos})...")

    # Clica no início da lista para dar foco e selecionar a primeira posição
    verificar_failsafe()
    print("  [Automação] Clicando no início da lista de filiais...")
    clicar_botao("inicio_lista_filial")
    aguardar_com_failsafe(0.5)

    # Navega através de cliques repetidos no botão 'navegar_lista_filial'
    if passos > 0:
        print(f"  [Automação] Navegando {passos} posição(ões) clicando no botão navegar_lista_filial...")
        for _ in range(passos):
            verificar_failsafe()
            clicar_botao("navegar_lista_filial")
            aguardar_com_failsafe(DELAYS.get("click", 0.5))
    elif passos < 0:
        print(f"  [Automação] AVISO: Navegação para cima não mapeada com cliques. Posição {passos}.")

    aguardar_com_failsafe(1.0)

    # Confirma a seleção com a tecla Enter
    if pressionar_enter:
        verificar_failsafe()
        print("  [Automação] Pressionando Enter para confirmar seleção da filial...")
        pyautogui.press("enter")
        aguardar_com_failsafe(DELAYS.get("enter", 1.5))

    print(f"  [Automação] Filial {codigo_str} selecionada com sucesso.")

def navegar_para_tela_titulos():
    print("\n  [Automação] Navegando pelos menus até a tela principal de Títulos...")
    botoes_sequencia = [
        "maleta",
        "financas",
        "contas_receber",
        "entrada_manutencao",
        "titulos_manutencao"
    ]
    
    for idx, btn in enumerate(botoes_sequencia, start=1):
        verificar_failsafe()
        info = obter_coordenadas_botao(btn)
        print(f"  [Automação] Menu ({idx}/5): {info.get('descricao')} [{btn}]")
        
        if btn == "titulos_manutencao":
            # Aguarda a animação do menu anterior terminar de expandir
            aguardar_com_failsafe(0.5)
            # Nós finais de TreeView (onde a tela efetivamente abre) geralmente exigem Duplo Clique no Senior ERP
            clicar_botao(btn, double_click=True)
        else:
            clicar_botao(btn)
            
        aguardar_com_failsafe(DELAYS.get("passo_menu", 1.0))

    aguardar_com_failsafe(DELAYS.get("abrir_janela", 2.0))
    print("  [Automação] Tela principal de Títulos/Manutenção alcançada com sucesso.")

def preencher_ate_data_entrada(titulo):
    """
    Preenche os campos iniciais do formulário F301TCR no Senior ERP:
    1. Nº Título -> Tab
    2. Tipo Título -> Tab
    3. Transação -> Tab
    4. Data Emissão -> Tab
    5. Data Entrada -> Tab
    """
    delay_campo = DELAYS.get("preenchimento", 0.2)

    print("\n  [Formulário ERP] Preenchendo campos até 'Data Entrada'...")

    # 1. Nº Título
    verificar_failsafe()
    num_titulo = str(titulo.get("numero_titulo", "")).strip()
    print(f"    ➜ [1/19] Nº Título: {num_titulo}")
    pyautogui.write(num_titulo, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)

    # 2. Tipo Título
    verificar_failsafe()
    tipo_titulo = str(titulo.get("tipo_titulo", "")).strip()
    print(f"    ➜ [2/19] Tipo Título: {tipo_titulo}")
    pyautogui.write(tipo_titulo, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)

    # 3. Transação
    verificar_failsafe()
    transacao = str(titulo.get("transacao", "")).strip()
    print(f"    ➜ [3/19] Transação: {transacao}")
    pyautogui.write(transacao, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)

    # 4. Data Emissão (formato DDMMYYYY)
    verificar_failsafe()
    data_emissao_br = formatar_data_br(titulo.get("data_emissao"), com_barras=False)
    print(f"    ➜ [4/19] Data Emissão: {data_emissao_br}")
    pyautogui.write(data_emissao_br, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)

    # 5. Data Entrada (formato DDMMYYYY)
    verificar_failsafe()
    data_entrada_br = formatar_data_br(titulo.get("data_entrada"), com_barras=False)
    print(f"    ➜ [5/19] Data Entrada: {data_entrada_br}")
    pyautogui.write(data_entrada_br, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    aguardar_com_failsafe(0.5)

    # Campo Cliente: fluxo completo de pesquisa
    verificar_failsafe()
    print("    ➜ [6/19] Clicando no botão de Pesquisar Cliente...")
    clicar_botao("pesquisa_cliente")
    aguardar_com_failsafe(DELAYS.get("abrir_janela", 2.0))

    # Alt+A para ativar/focar o campo "Valor" da pesquisa
    verificar_failsafe()
    nome_cliente = str(titulo.get("cliente", "")).strip()
    print(f"    ➜ [7/19] Alt+A → Digitando nome do cliente: {nome_cliente}")
    pyautogui.hotkey("alt", "a")
    aguardar_com_failsafe(0.5)

    # Digita o nome do cliente via clipboard (suporta caracteres especiais: ç, ã, etc.)
    verificar_failsafe()
    digitar_texto(nome_cliente)
    aguardar_com_failsafe(0.3)

    # Clica no botão filtrar_cliente para iniciar a pesquisa
    verificar_failsafe()
    print("    ➜ [8/19] Clicando em Filtrar Cliente...")
    clicar_botao("filtrar_cliente")
    aguardar_com_failsafe(DELAYS.get("abrir_janela", 2.0))

    # Verifica se a pesquisa retornou mais de 1 cliente com o mesmo nome
    if verificar_cliente_duplicado(nome_cliente):
        log(f"  ➜ [CLIENTE DUPLICADO] Múltiplos clientes encontrados para '{nome_cliente}'. Pulando este título!", "AVISO")
        verificar_failsafe()
        log("  [Automação] Clicando em cancelar_filtro_cliente...")
        clicar_botao("cancelar_filtro_cliente")
        aguardar_com_failsafe(1.0)
        log("  [Automação] Clicando em cancelar_titulo...")
        clicar_botao("cancelar_titulo")
        pyautogui.press("enter")
        aguardar_com_failsafe(1.0)
        return False  # Retorna False indicando que o título foi pulado por duplicidade

    # Duplo clique no botão confirmar_cliente para selecionar o primeiro resultado
    verificar_failsafe()
    print("    ➜ [9/19] Duplo clique em Confirmar Cliente...")
    clicar_botao("confirmar_cliente")
    pyautogui.press("enter")
    aguardar_com_failsafe(DELAYS.get("click", 1.0))

    # 4x Tab para chegar ao campo de Observação
    verificar_failsafe()
    print("    ➜ [10/19] Navegando até campo Observação (4x Tab)...")
    for _ in range(4):
        pyautogui.press("tab")
        aguardar_com_failsafe(0.2)

    # Campo Observação: digita via clipboard (suporta caracteres especiais)
    verificar_failsafe()
    observacao = str(titulo.get("observacao", "")).strip()
    print(f"    ➜ [11/19] Digitando Observação: {observacao}")
    digitar_texto(observacao)
    aguardar_com_failsafe(delay_campo)

    # Pula o campo Natureza com 1x Tab
    verificar_failsafe()
    print("    ➜ [12/19] Pulando campo Natureza (1x Tab)...")
    pyautogui.press("tab")
    pyautogui.press("tab")
    aguardar_com_failsafe(0.2)

    # Campo Projeto
    verificar_failsafe()
    projeto = str(titulo.get("projeto", "")).strip()
    print(f"    ➜ [13/19] Digitando Projeto: {projeto}")
    pyautogui.write(projeto, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)

    # Campo Fase
    verificar_failsafe()
    fase = str(titulo.get("fase", "")).strip()
    print(f"    ➜ [14/19] Digitando Fase: {fase}")
    pyautogui.write(fase, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)

    # Campo Conta Financeira
    verificar_failsafe()
    conta_financeira = str(titulo.get("conta_financeira", "")).strip()
    print(f"    ➜ [15/19] Digitando Conta Financeira: {conta_financeira}")
    pyautogui.write(conta_financeira, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)

    # Campo Centro de Custo (código)
    verificar_failsafe()
    cod_centro_custo = str(titulo.get("cod._do_centro_de_custo", "")).strip()
    print(f"    ➜ [16/19] Digitando Centro de Custo: {cod_centro_custo}")
    pyautogui.write(cod_centro_custo, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)

    # Campo Data de Vencimento (formato DDMMYYYY)
    verificar_failsafe()
    data_vencimento_br = formatar_data_br(titulo.get("data_vencimento"), com_barras=False, inverter_dia_mes=True)
    print(f"    ➜ [17/19] Digitando Data de Vencimento: {data_vencimento_br}")
    pyautogui.write(data_vencimento_br, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)

    # Campo Valor Original
    verificar_failsafe()
    valor_original = str(titulo.get("valor_original", "")).strip().replace(".", ",")
    print(f"    ➜ [18/19] Digitando Valor Original: {valor_original}")
    pyautogui.write(valor_original, interval=0.03)
    aguardar_com_failsafe(delay_campo)
    pyautogui.press("tab")
    aguardar_com_failsafe(delay_campo)
    # 34x Tab para percorrer os campos restantes e efetivar a gravação sem avançar para o 2º campo do novo título
    verificar_failsafe()
    print("    ➜ [19/19] Finalizando cadastro (34x Tab)...")
    for _ in range(34):
        pyautogui.press("tab")
        aguardar_com_failsafe(0.1)

    print("  ✔ [Formulário ERP] Cadastro do título finalizado com sucesso!")
    return True

def executar_automacao_completa(modo_teste=False):
    log("=======================================================")
    log("      AUTOMAÇÃO DE CRIAÇÃO DE TÍTULOS - SENIOR ERP     ")
    log("=======================================================")

    titulos = carregar_titulos()
    if not titulos:
        log("Nenhum título para processar.", "AVISO")
        return

    total_titulos = len(titulos)
    log(f"Total de títulos a processar: {total_titulos}")

    # Verificar de onde retomar (mecanismo de checkpoint)
    ultimo_indice_concluido = obter_ultimo_indice_processado()
    if ultimo_indice_concluido > 0 and ultimo_indice_concluido < total_titulos:
        log(f"RETOMADA ATIVADA: Continuando a partir do título {ultimo_indice_concluido + 1} de {total_titulos}.", "AVISO")
    elif ultimo_indice_concluido >= total_titulos:
        log("Todos os títulos do arquivo já foram marcados como concluídos. Reiniciando do título 1...", "AVISO")
        ultimo_indice_concluido = 0

    try:
        # 1. Contagem regressiva de 5 segundos
        iniciar_contagem_regressiva(5)

        filial_atual = None
        tela_aberta = False  # Indica se existe tela (ex: F301TCR) aberta no ERP

        # 2. Iteração pelos títulos a partir do ponto de parada
        for index, titulo in enumerate(titulos, start=1):
            if index <= ultimo_indice_concluido:
                # Pula os títulos que já foram processados anteriormente
                continue

            verificar_failsafe()

            num_titulo = titulo.get("numero_titulo")
            cliente = titulo.get("cliente")
            projeto_filial = str(titulo.get("projeto")).strip()
            valor = titulo.get("valor_original")
            vencimento = titulo.get("data_vencimento")
            obs = titulo.get("observacao")
            centro_custo = titulo.get("centro_custo")

            log(f"-------------------------------------------------------")
            log(f"[Título {index}/{total_titulos}] N° Título: {num_titulo}")
            log(f"  • Cliente: {cliente}")
            log(f"  • Filial (Projeto): {projeto_filial}")
            log(f"  • Valor: R$ {valor}")
            log(f"  • Vencimento: {vencimento}")
            log(f"  • Centro de Custo: {centro_custo}")
            log(f"  • Observação: {obs}")

            # Troca de filial se necessário
            if filial_atual != projeto_filial:
                log(f"  ➜ [TROCA DE FILIAL] Filial anterior: '{filial_atual}' → Nova Filial: '{projeto_filial}'")
                abrir_tela_trocar_filial(tem_tela_aberta=tela_aberta)
                selecionar_filial(projeto_filial)
                navegar_para_tela_titulos()
                filial_atual = projeto_filial
                tela_aberta = True  # Tela F301TCR passa a estar aberta
            else:
                log(f"  ➜ [MANTER FILIAL] Filial atual já é '{filial_atual}'. Mantendo formulário F301TCR pronto.")
                if not tela_aberta:
                    clicar_botao("titulos_manutencao", double_click=True)
                    tela_aberta = True

            # Preencher campos do formulário F301TCR
            sucesso = preencher_ate_data_entrada(titulo)

            if not sucesso:
                # O título foi pulado por duplicidade de cliente
                salvar_progresso(index, num_titulo)
                registrar_titulo_planilha(titulo, index, status="PULADO - CLIENTE DUPLICADO")
                registrar_titulo_pulado(titulo, index, motivo="Cliente Duplicado")
                log(f"  ✔ Título {num_titulo} registrado como PULADO nas planilhas com sucesso.", "AVISO")
                aguardar_com_failsafe(DELAYS.get("entre_titulos", 1.5))
                continue

            # Registrar sucesso do título na planilha e no checkpoint de progresso
            salvar_progresso(index, num_titulo)
            registrar_titulo_planilha(titulo, index, status="SUCESSO")

            # Verifica se o próximo título é de filial diferente
            proximo_titulo = titulos[index] if index < len(titulos) else None  # index já começa em 1
            proxima_filial = str(proximo_titulo.get("projeto", "")).strip() if proximo_titulo else None

            if proxima_filial and proxima_filial != projeto_filial:
                # Clica em Sair para fechar a tela sem popup de confirmação de troca de filial
                verificar_failsafe()
                log(f"  [Automação] Próximo título é da filial '{proxima_filial}'. Clicando em Sair...")
                clicar_botao("sair_titulo")
                aguardar_com_failsafe(DELAYS.get("abrir_janela", 2.0))
                tela_aberta = False

            aguardar_com_failsafe(DELAYS.get("entre_titulos", 1.5))

        # Se concluiu todos com sucesso
        limpar_progresso()
        log("=======================================================")
        log(f"[Automação] Processamento concluído com sucesso!")
        log(f"[Automação] Total de {total_titulos} títulos processados.")
        log(f"[Automação] Arquivos salvos na pasta 'outputs':")
        log(f"  - Logs: {LOG_TXT_PATH.name}")
        log(f"  - Planilha: {EXCEL_OUTPUT_PATH.name if EXCEL_DISPONIVEL else CSV_OUTPUT_PATH.name}")
        log("=======================================================")

    except pyautogui.FailSafeException:
        log("INTERRUPÇÃO DE EMERGÊNCIA DISPARADA! Mouse no canto superior esquerdo (Fail-Safe).", "INTERRUPÇÃO")
        print("\n\n" + "!" * 65)
        print(" INTERRUPÇÃO DE EMERGÊNCIA DISPARADA! ")
        print(" Mouse movido para o canto superior esquerdo (Fail-Safe).")
        print(" A automação foi PARADA IMEDIATAMENTE por segurança.")
        print(f" Progresso salvo no título {obter_ultimo_indice_processado()}. Execute novamente para continuar de onde parou.")
        print("!" * 65 + "\n")
    except KeyboardInterrupt:
        log("Interrompido pelo usuário via teclado (Ctrl+C).", "INTERRUPÇÃO")
        print(f"\n\n[Automação] Interrompido pelo usuário via teclado. Progresso salvo no título {obter_ultimo_indice_processado()}.")
    except Exception as e:
        log(f"Erro inesperado durante a execução: {e}", "ERRO")
        raise e

if __name__ == "__main__":
    executar_automacao_completa()
